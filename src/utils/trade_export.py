"""
Export trade journal (data/trades.json) to a formatted .xlsx workbook.

Produces four sheets:
  - Trades     : flat per-trade view (entries + exits + P&L)
  - Summary    : aggregate stats (WR, PF, gross/net, fees est.)
  - By Symbol  : per-symbol breakdown
  - Daily P&L  : UTC-day net P&L + cumulative equity curve

Designed for monitoring the live test — open in Excel/LibreOffice for the
weekly review in LIVE_TEST.md. All totals use Excel formulas so the workbook
stays reactive if rows get edited by hand.
"""
from __future__ import annotations

import json
import os
from collections import defaultdict
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.utils.logger import logger

JOURNAL_PATH = os.path.join(os.getcwd(), "data", "trades.json")

# ---- Styling constants (per xlsx skill: blue inputs, black formulas, green cross-sheet)
FONT_NAME = "Arial"

_HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", start_color="1F4E78")  # dark blue
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

_INPUT_FONT = Font(name=FONT_NAME, color="0000FF")          # blue = hardcoded input
_FORMULA_FONT = Font(name=FONT_NAME, color="000000")        # black = formula
_CROSS_FONT = Font(name=FONT_NAME, color="008000")          # green = cross-sheet link

_WIN_FILL = PatternFill("solid", start_color="E2F0D9")      # soft green
_LOSS_FILL = PatternFill("solid", start_color="FCE4D6")     # soft red
_KEY_FILL = PatternFill("solid", start_color="FFFF00")      # yellow = key metric

_THIN = Side(border_style="thin", color="BFBFBF")
_BORDER = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)

# Number formats
_FMT_CCY = '"$"#,##0.00;("$"#,##0.00);-'
_FMT_PCT = '0.0%;(0.0%);-'
_FMT_NUM = "#,##0.00;(#,##0.00);-"
_FMT_INT = "#,##0;(#,##0);-"
_FMT_DATE = "yyyy-mm-dd hh:mm"


def _load_trades() -> list[dict]:
    if not os.path.exists(JOURNAL_PATH):
        return []
    try:
        with open(JOURNAL_PATH, "r") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        logger.warning(f"trade_export: cannot load journal: {e}")
        return []


def _parse_ts(s: Optional[str]):
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except (ValueError, TypeError):
        return None


def _write_header(ws, row: int, columns: list[str]):
    for idx, title in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=idx, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER


def _autosize(ws, min_width: int = 10, max_width: int = 40):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        width = min_width
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) + 2 > width:
                width = len(s) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, width)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

_TRADE_COLS = [
    "ID", "Opened (UTC)", "Closed (UTC)", "Symbol", "Side",
    "Entry", "Exit", "Amount",
    "Notional ($)", "PnL ($)", "PnL %", "Confidence",
    "Status", "Sentiment", "Reasoning",
]


def _write_trades(wb: Workbook, trades: list[dict]) -> int:
    """Write the flat Trades sheet. Returns number of trade rows."""
    ws = wb.active
    ws.title = "Trades"

    _write_header(ws, 1, _TRADE_COLS)

    # Sort oldest → newest so the running curve on "Daily P&L" is monotonic
    ordered = sorted(
        trades,
        key=lambda t: _parse_ts(t.get("timestamp")) or datetime.min,
    )

    r = 2
    for t in ordered:
        opened = _parse_ts(t.get("timestamp"))
        closed = _parse_ts(t.get("exit_time"))
        entry = t.get("entry_price")
        exit_ = t.get("exit_price")
        amount = t.get("amount")
        pnl = t.get("pnl")
        pnl_pct = t.get("pnl_pct")
        status = t.get("status", "")
        sentiment = t.get("sentiment", "")
        reasoning = (t.get("reasoning") or "")[:200]

        ws.cell(row=r, column=1, value=t.get("id", ""))
        ws.cell(row=r, column=2, value=opened).number_format = _FMT_DATE
        ws.cell(row=r, column=3, value=closed).number_format = _FMT_DATE
        ws.cell(row=r, column=4, value=t.get("symbol", ""))
        ws.cell(row=r, column=5, value=t.get("side", ""))
        ws.cell(row=r, column=6, value=entry).number_format = _FMT_NUM
        ws.cell(row=r, column=7, value=exit_).number_format = _FMT_NUM
        ws.cell(row=r, column=8, value=amount).number_format = "0.000000"

        # Notional = entry * amount (use Excel formula so edits propagate)
        ws.cell(row=r, column=9, value=f"=IFERROR(F{r}*H{r},0)").number_format = _FMT_CCY
        pnl_cell = ws.cell(row=r, column=10, value=pnl)
        pnl_cell.number_format = _FMT_CCY
        pct_cell = ws.cell(row=r, column=11, value=(pnl_pct / 100) if isinstance(pnl_pct, (int, float)) else None)
        pct_cell.number_format = _FMT_PCT

        conf = t.get("confidence")
        ws.cell(row=r, column=12, value=conf).number_format = "0.00"
        ws.cell(row=r, column=13, value=status)
        ws.cell(row=r, column=14, value=sentiment)
        ws.cell(row=r, column=15, value=reasoning).alignment = Alignment(wrap_text=True, vertical="top")

        # Highlight the row based on outcome
        if isinstance(pnl, (int, float)):
            fill = _WIN_FILL if pnl > 0 else (_LOSS_FILL if pnl < 0 else None)
            if fill is not None:
                for col in range(1, len(_TRADE_COLS) + 1):
                    ws.cell(row=r, column=col).fill = fill

        r += 1

    ws.freeze_panes = "A2"
    _autosize(ws)
    # Reasoning column gets generous width, plus enable wrap
    ws.column_dimensions["O"].width = 60
    return r - 2  # number of data rows written


def _write_summary(wb: Workbook, n_trades: int):
    """Summary sheet — formulas pull from Trades so edits propagate."""
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    _write_header(ws, 1, ["Metric", "Value"])

    # Range expressions across the Trades sheet
    pnl_range = f"Trades!J2:J{n_trades + 1}"
    status_range = f"Trades!M2:M{n_trades + 1}"
    pct_range = f"Trades!K2:K{n_trades + 1}"
    notional_range = f"Trades!I2:I{n_trades + 1}"

    rows = [
        ("Generated (UTC)", datetime.utcnow().strftime("%Y-%m-%d %H:%M"), None, False),
        ("Source rows (all)", n_trades, _FMT_INT, False),
        # Totals
        ("Closed trades",
         f'=COUNTIF({status_range},"closed")', _FMT_INT, False),
        ("Open trades",
         f'=COUNTIF({status_range},"open")', _FMT_INT, False),
        ("Wins (pnl > 0)",
         f'=COUNTIFS({status_range},"closed",{pnl_range},">0")', _FMT_INT, False),
        ("Losses (pnl <= 0)",
         f'=COUNTIFS({status_range},"closed",{pnl_range},"<=0") - '
         f'COUNTIFS({status_range},"closed",{pnl_range},"=0")', _FMT_INT, False),
        ("Breakeven (pnl = 0)",
         f'=COUNTIFS({status_range},"closed",{pnl_range},"=0")', _FMT_INT, False),
        # Win rate
        ("Win rate",
         f'=IFERROR(COUNTIFS({status_range},"closed",{pnl_range},">0")/'
         f'COUNTIF({status_range},"closed"),0)', _FMT_PCT, True),
        # Gross / net
        ("Gross profit",
         f'=SUMIFS({pnl_range},{status_range},"closed",{pnl_range},">0")', _FMT_CCY, False),
        ("Gross loss",
         f'=SUMIFS({pnl_range},{status_range},"closed",{pnl_range},"<=0")', _FMT_CCY, False),
        ("Net P&L",
         f'=SUMIFS({pnl_range},{status_range},"closed")', _FMT_CCY, True),
        # Averages
        ("Average win",
         f'=IFERROR(AVERAGEIFS({pnl_range},{status_range},"closed",{pnl_range},">0"),0)', _FMT_CCY, False),
        ("Average loss",
         f'=IFERROR(AVERAGEIFS({pnl_range},{status_range},"closed",{pnl_range},"<=0"),0)', _FMT_CCY, False),
        ("Average P&L %",
         f'=IFERROR(AVERAGE({pct_range}),0)', _FMT_PCT, False),
        # Profit factor = gross profit / |gross loss|
        ("Profit factor",
         f'=IFERROR(SUMIFS({pnl_range},{status_range},"closed",{pnl_range},">0") / '
         f'ABS(SUMIFS({pnl_range},{status_range},"closed",{pnl_range},"<0")),0)', "0.00", True),
        # Best / worst
        ("Best trade",
         f'=IFERROR(MAX({pnl_range}),0)', _FMT_CCY, False),
        ("Worst trade",
         f'=IFERROR(MIN({pnl_range}),0)', _FMT_CCY, False),
        # Notional stats
        ("Total notional traded",
         f'=IFERROR(SUM({notional_range}),0)', _FMT_CCY, False),
        ("Avg notional / trade",
         f'=IFERROR(AVERAGE({notional_range}),0)', _FMT_CCY, False),
    ]

    for i, (label, value, fmt, highlight) in enumerate(rows, start=2):
        label_cell = ws.cell(row=i, column=1, value=label)
        value_cell = ws.cell(row=i, column=2, value=value)
        label_cell.font = Font(name=FONT_NAME, bold=True)
        label_cell.border = _BORDER
        value_cell.border = _BORDER
        if isinstance(value, str) and value.startswith("="):
            value_cell.font = _CROSS_FONT
        else:
            value_cell.font = _INPUT_FONT
        if fmt:
            value_cell.number_format = fmt
        if highlight:
            value_cell.fill = _KEY_FILL

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20


def _write_by_symbol(wb: Workbook, trades: list[dict], n_trades: int):
    ws = wb.create_sheet("By Symbol")
    headers = ["Symbol", "Trades", "Wins", "Losses", "Win Rate", "Net P&L", "Avg P&L"]
    _write_header(ws, 1, headers)

    symbols = sorted({t.get("symbol", "") for t in trades if t.get("symbol")})
    if not symbols:
        ws.cell(row=2, column=1, value="(no trades)")
        ws.column_dimensions["A"].width = 16
        return

    pnl_range = f"Trades!J2:J{n_trades + 1}"
    status_range = f"Trades!M2:M{n_trades + 1}"
    symbol_range = f"Trades!D2:D{n_trades + 1}"

    for r, sym in enumerate(symbols, start=2):
        ws.cell(row=r, column=1, value=sym).font = _INPUT_FONT
        # Total trades for symbol (any status)
        ws.cell(row=r, column=2,
                value=f'=COUNTIF({symbol_range},A{r})').number_format = _FMT_INT
        # Wins
        ws.cell(row=r, column=3,
                value=f'=COUNTIFS({symbol_range},A{r},{status_range},"closed",{pnl_range},">0")'
                ).number_format = _FMT_INT
        # Losses
        ws.cell(row=r, column=4,
                value=f'=COUNTIFS({symbol_range},A{r},{status_range},"closed",{pnl_range},"<=0")'
                ).number_format = _FMT_INT
        # WR
        ws.cell(row=r, column=5,
                value=f'=IFERROR(C{r}/(C{r}+D{r}),0)').number_format = _FMT_PCT
        # Net PnL
        ws.cell(row=r, column=6,
                value=f'=SUMIFS({pnl_range},{symbol_range},A{r},{status_range},"closed")'
                ).number_format = _FMT_CCY
        # Avg PnL
        ws.cell(row=r, column=7,
                value=f'=IFERROR(F{r}/(C{r}+D{r}),0)').number_format = _FMT_CCY

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = _BORDER

    ws.column_dimensions["A"].width = 14
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 14


def _write_daily_pnl(wb: Workbook, trades: list[dict]):
    ws = wb.create_sheet("Daily P&L")
    headers = ["Date (UTC)", "Trades Closed", "Wins", "Losses", "Net P&L", "Cumulative"]
    _write_header(ws, 1, headers)

    buckets: dict[str, dict] = defaultdict(lambda: {"n": 0, "w": 0, "l": 0, "pnl": 0.0})
    for t in trades:
        if t.get("status") != "closed":
            continue
        ts = _parse_ts(t.get("exit_time")) or _parse_ts(t.get("timestamp"))
        if ts is None:
            continue
        pnl = t.get("pnl")
        if pnl is None:
            continue
        key = ts.strftime("%Y-%m-%d")
        b = buckets[key]
        b["n"] += 1
        b["pnl"] += float(pnl)
        if pnl > 0:
            b["w"] += 1
        else:
            b["l"] += 1

    if not buckets:
        ws.cell(row=2, column=1, value="(no closed trades yet)")
        ws.column_dimensions["A"].width = 22
        return

    dates = sorted(buckets.keys())
    for r, date in enumerate(dates, start=2):
        b = buckets[date]
        ws.cell(row=r, column=1, value=date).font = _INPUT_FONT
        ws.cell(row=r, column=2, value=b["n"]).number_format = _FMT_INT
        ws.cell(row=r, column=3, value=b["w"]).number_format = _FMT_INT
        ws.cell(row=r, column=4, value=b["l"]).number_format = _FMT_INT
        pnl_cell = ws.cell(row=r, column=5, value=round(b["pnl"], 4))
        pnl_cell.number_format = _FMT_CCY
        if b["pnl"] > 0:
            pnl_cell.fill = _WIN_FILL
        elif b["pnl"] < 0:
            pnl_cell.fill = _LOSS_FILL
        # Cumulative: formula = prev cumulative + today's net
        if r == 2:
            cum_formula = f"=E{r}"
        else:
            cum_formula = f"=F{r - 1}+E{r}"
        ws.cell(row=r, column=6, value=cum_formula).number_format = _FMT_CCY

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = _BORDER

    ws.column_dimensions["A"].width = 14
    for col in "BCDEF":
        ws.column_dimensions[col].width = 14
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Public entrypoint
# ---------------------------------------------------------------------------

def export_trades_to_xlsx(output_path: str, days: Optional[int] = None) -> dict:
    """
    Export the trade journal to an xlsx workbook.

    Args:
        output_path: where to write the .xlsx file.
        days: optional — include only trades whose opening timestamp is within
              the last N days. None = include everything.

    Returns:
        Stats dict: {"rows": int, "path": str, "closed": int, "open": int}
    """
    trades = _load_trades()

    if days is not None and days > 0:
        from datetime import timedelta
        cutoff = datetime.utcnow() - timedelta(days=days)
        filtered: list[dict] = []
        for t in trades:
            ts = _parse_ts(t.get("timestamp"))
            if ts is None or ts.replace(tzinfo=None) >= cutoff:
                filtered.append(t)
        trades = filtered

    wb = Workbook()
    n_rows = _write_trades(wb, trades)
    _write_summary(wb, n_rows)
    _write_by_symbol(wb, trades, n_rows)
    _write_daily_pnl(wb, trades)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)

    closed = sum(1 for t in trades if t.get("status") == "closed")
    open_ = sum(1 for t in trades if t.get("status") == "open")
    stats = {
        "rows": n_rows,
        "path": os.path.abspath(output_path),
        "closed": closed,
        "open": open_,
    }
    logger.info(
        f"trade_export: wrote {n_rows} rows ({closed} closed, {open_} open) → {stats['path']}"
    )
    return stats
